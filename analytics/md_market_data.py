# analytics/md_market_data.py
"""Maryland Cannabis Market Data Analysis.

Parses official MCA (Maryland Cannabis Administration) data dashboard exports
to provide market share estimates, sales trends, and category breakdowns.
"""

import os
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
from dataclasses import dataclass, field

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class MarketSummary:
    """Summary of Maryland cannabis market data."""
    total_sales: float = 0.0
    sales_by_type: Dict[str, float] = field(default_factory=dict)
    sales_by_category: Dict[str, float] = field(default_factory=dict)
    sales_by_month: Dict[str, float] = field(default_factory=dict)
    sales_by_customer_type: Dict[str, float] = field(default_factory=dict)
    dispensary_count: int = 0
    dispensaries_by_county: Dict[str, int] = field(default_factory=dict)
    year: int = 2024

    def get_category_share(self, category: str) -> float:
        """Get market share percentage for a category."""
        if self.total_sales == 0:
            return 0.0
        return (self.sales_by_type.get(category, 0) / self.total_sales) * 100


@dataclass
class Dispensary:
    """Maryland dispensary from official state data."""
    license_number: str
    business_name: str
    location_name: str
    trade_name: str
    county: str
    address: str
    region: str
    lat: Optional[float] = None
    lng: Optional[float] = None
    url: Optional[str] = None
    logo_url: Optional[str] = None


class MDMarketAnalyzer:
    """Analyzes Maryland cannabis market data from MCA exports."""

    DEFAULT_DATA_PATH = "data/md_sales_data.xlsx"

    def __init__(self, data_path: Optional[str] = None):
        """Initialize with path to MCA data export.

        Args:
            data_path: Path to Excel file from MCA data dashboard
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required: pip install openpyxl")

        self.data_path = data_path or self.DEFAULT_DATA_PATH
        self._workbook = None
        self._dispensaries: List[Dispensary] = []
        self._market_summary: Optional[MarketSummary] = None

    def _load_workbook(self):
        """Load the Excel workbook."""
        if self._workbook is None:
            self._workbook = load_workbook(self.data_path, data_only=True)
        return self._workbook

    def get_dispensaries(self) -> List[Dispensary]:
        """Get list of all licensed dispensaries from state data."""
        if self._dispensaries:
            return self._dispensaries

        wb = self._load_workbook()
        ws = wb['Dispensaries Address Data']

        for row in ws.iter_rows(min_row=2, values_only=True):
            lat, lng, license_num, logo, url, business, location, trade, county, address, region = row
            if license_num:
                self._dispensaries.append(Dispensary(
                    license_number=license_num,
                    business_name=business or "",
                    location_name=location or "",
                    trade_name=trade or "",
                    county=county or "",
                    address=address or "",
                    region=region or "",
                    lat=float(lat) if lat else None,
                    lng=float(lng) if lng else None,
                    url=url,
                    logo_url=logo
                ))

        return self._dispensaries

    def get_dispensaries_by_county(self) -> Dict[str, int]:
        """Get dispensary count by county."""
        wb = self._load_workbook()
        ws = wb['Dispensaries by County']

        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            county, count = row
            if county and count is not None:
                result[county] = int(count)

        return result

    def get_market_summary(self, year: int = 2024) -> MarketSummary:
        """Get market summary for a specific year.

        Args:
            year: Year to summarize (default 2024)

        Returns:
            MarketSummary with sales breakdowns
        """
        wb = self._load_workbook()
        ws = wb['Market Sales Data']

        summary = MarketSummary(year=year)

        for row in ws.iter_rows(min_row=2, values_only=True):
            product_type, category, row_year, month, customer_type, total_price = row

            if row_year != year or not total_price:
                continue

            summary.total_sales += total_price

            # By product type (Flower, Concentrate, Edible, Other)
            if product_type:
                summary.sales_by_type[product_type] = summary.sales_by_type.get(product_type, 0) + total_price

            # By detailed category
            if category:
                summary.sales_by_category[category] = summary.sales_by_category.get(category, 0) + total_price

            # By month
            if month:
                summary.sales_by_month[month] = summary.sales_by_month.get(month, 0) + total_price

            # By customer type (Medical vs Adult-Use)
            if customer_type:
                summary.sales_by_customer_type[customer_type] = summary.sales_by_customer_type.get(customer_type, 0) + total_price

        # Add dispensary data
        summary.dispensaries_by_county = self.get_dispensaries_by_county()
        summary.dispensary_count = sum(summary.dispensaries_by_county.values())

        self._market_summary = summary
        return summary

    def get_monthly_sales(self, year: int = 2024) -> List[Tuple[str, float]]:
        """Get monthly sales totals for a year.

        Returns:
            List of (month_name, total_sales) tuples
        """
        summary = self.get_market_summary(year)

        # Order months correctly
        month_order = [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ]

        return [(m, summary.sales_by_month.get(m, 0)) for m in month_order]

    def get_category_breakdown(self, year: int = 2024) -> List[Dict]:
        """Get sales breakdown by product type with percentages.

        Returns:
            List of dicts with category, sales, and percentage
        """
        summary = self.get_market_summary(year)

        result = []
        for category, sales in sorted(summary.sales_by_type.items(), key=lambda x: -x[1]):
            result.append({
                'category': category,
                'sales': sales,
                'percentage': (sales / summary.total_sales * 100) if summary.total_sales > 0 else 0
            })

        return result

    def estimate_dispensary_market_share(
        self,
        dispensary_name: str,
        county: str,
        estimated_monthly_revenue: float
    ) -> Dict:
        """Estimate a dispensary's market share.

        Args:
            dispensary_name: Name of the dispensary
            county: County where dispensary is located
            estimated_monthly_revenue: Estimated monthly revenue

        Returns:
            Dict with market share estimates
        """
        summary = self.get_market_summary()

        # State-level share
        annual_revenue = estimated_monthly_revenue * 12
        state_share = (annual_revenue / summary.total_sales * 100) if summary.total_sales > 0 else 0

        # County-level estimate (proportional to county dispensary count)
        county_dispensaries = summary.dispensaries_by_county.get(county, 1)
        county_share = (1 / county_dispensaries * 100) if county_dispensaries > 0 else 0

        # Average revenue per dispensary
        avg_revenue = summary.total_sales / summary.dispensary_count if summary.dispensary_count > 0 else 0

        return {
            'dispensary': dispensary_name,
            'county': county,
            'estimated_annual_revenue': annual_revenue,
            'state_market_share': state_share,
            'county_dispensaries': county_dispensaries,
            'county_market_share_estimate': county_share,
            'state_avg_revenue_per_dispensary': avg_revenue,
            'performance_vs_avg': (annual_revenue / avg_revenue * 100) if avg_revenue > 0 else 0
        }

    def print_summary(self, year: int = 2024):
        """Print a formatted market summary."""
        summary = self.get_market_summary(year)

        print(f"\n{'='*60}")
        print(f"MARYLAND CANNABIS MARKET SUMMARY - {year}")
        print(f"{'='*60}\n")

        print(f"Total Market Sales: ${summary.total_sales:,.2f}")
        print(f"Total Licensed Dispensaries: {summary.dispensary_count}")
        print(f"Average Revenue per Dispensary: ${summary.total_sales / summary.dispensary_count:,.2f}")

        print(f"\n{'--- Sales by Product Type ---':^60}")
        for category in sorted(summary.sales_by_type.keys(), key=lambda x: -summary.sales_by_type[x]):
            sales = summary.sales_by_type[category]
            pct = summary.get_category_share(category)
            print(f"  {category:30} ${sales:>15,.2f}  ({pct:5.1f}%)")

        print(f"\n{'--- Sales by Customer Type ---':^60}")
        for ctype, sales in sorted(summary.sales_by_customer_type.items(), key=lambda x: -x[1]):
            pct = (sales / summary.total_sales * 100) if summary.total_sales > 0 else 0
            print(f"  {ctype:30} ${sales:>15,.2f}  ({pct:5.1f}%)")

        print(f"\n{'--- Dispensaries by County (Top 10) ---':^60}")
        for county, count in sorted(summary.dispensaries_by_county.items(), key=lambda x: -x[1])[:10]:
            print(f"  {county:30} {count:>3} dispensaries")

        print(f"\n{'='*60}\n")


def main():
    """Run market analysis."""
    analyzer = MDMarketAnalyzer()
    analyzer.print_summary(2024)

    # Show monthly trend
    print("\n--- Monthly Sales 2024 ---")
    for month, sales in analyzer.get_monthly_sales(2024):
        print(f"  {month:12} ${sales:>15,.2f}")


if __name__ == "__main__":
    main()
